import csv

from bs4 import BeautifulSoup
import requests
import locale
import re

import json
import os
from docx import Document


test_url3 = f'https://soviet-trawler.narod.ru/main_r/front7_r.html'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0'
                    ' YaBrowser/23.3.0.2246 Yowser/2.5 Safari/537.36', 'accept': '*/*'}

req = requests.get(test_url3, headers=HEADERS, params=None)
src = req.text
soup = BeautifulSoup(src, 'lxml')
i = 0

links = []
links_date = {}
ship_data_all ={}
tbody = soup.find(class_= "tbody")

tbody = soup.find('tbody')
if tbody:
    for link in tbody.find_all('a'):
        links.append("https://soviet-trawler.narod.ru/main_r/" + link.get('href'))
    
for link in links:

    url_table_ship = link
    req = requests.get(url_table_ship, headers=HEADERS, params=None)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')
 


    # Находим все строки таблицы с классом shiplist
    table_rows = soup.select('table.shiplist tr')
    for row in table_rows[2:]:
        ship_data = []
        # Находим все ячейки в строке
        cells = row.find_all('td')
        photo_tag = cells[0].find('img')
        if photo_tag:
            photo = photo_tag.get('src')
            photo = 'https://soviet-trawler.narod.ru/' + photo
        else:
            photo = None  # Или другое значение по умолчанию
        # Извлекаем данные из ячеек

        ship_type = cells[1].get_text(strip=True).replace('\r','').replace('\n','')
        dimensions = cells[2].get_text(strip=True)
        build_years = cells[3].get_text(strip=True)
        build_countries = cells[4].get_text(strip=True)
        try:
            all_link_ships = "https://soviet-trawler.narod.ru/main_r/" + cells[1].find('a').get('href')
        except:
            pass
        # Создаем словарь с данными и добавляем его в список
        ship_info = {
            'photo':  photo,
            'Тип судна': ship_type,
            'Главные размерения': dimensions,
            'Годы постройки': build_years,
            'Страны постройки': build_countries
        }
        for label, value in ship_info.items():
            ship_data.append((label, value))
    
    # Печатаем результат
    # for ship in ship_data:
    #     print(ship)
    # print(all_link_ships)


        url_table_ship = all_link_ships
        req = requests.get(url_table_ship, headers=HEADERS, params=None)
        src = req.text
        soup = BeautifulSoup(src, 'lxml')

        # ship_info_data = []

        # Находим все строки таблицы с классом shipinfo
        table_rows = soup.select('table.shipinfo tr')

        # Пропускаем первые две строки, так как они содержат заголовки
        for row in table_rows[2:]:
            # Находим все ячейки в строке
            cells = row.find_all('td')

            # Извлекаем данные из ячеек
            if len(cells) == 2:
                label = cells[0].get_text(strip=True).replace('\r','').replace('\n','')
                value = cells[1].get_text(strip=True).replace('\r','').replace('\n','')
                # Добавляем данные в массив
                ship_data.append((label, value))
        ship_data_all[i] = ship_data
        i = i + 1
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

# Печатаем результат
# for label, value in ship_data:
#     print(f"{label}: {value}")
current_row = 1

# Записываем заголовки столбцов в первую строку
for row_number, data in ship_data_all.items():
    headers = [label for label, value in data]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col, value=header)
    current_row += 1
    break 

# Перебираем данные для каждого судна и записываем их в соответствующие строки и столбцы
for row_number, data in ship_data_all.items():
    current_row += 1  # Переходим на следующую строку
    for col, (label, value) in enumerate(data, start=1):
        ws.cell(row=current_row, column=col, value=value)

# Сохраняем книгу Excel
wb.save('ships_info.xlsx')


# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# # Заполняем таблицу данными
# for col, (label, value) in enumerate(ship_data, start=1):
#     ws.cell(row=1, column=col, value=label)
#     ws.cell(row=2, column=col, value=value)

# # Сохраняем книгу Excel
# wb.save('ships_info_transposed.xlsx')


# Создаем новую книгу Excel
# wb = Workbook()
# ws = wb.active

# # Заголовки столбцов
# headers = set()
# for ship in ship_data:
#     headers.update(ship.keys())

# # Заголовки столбцов
# headers = list(headers)
# for col, header in enumerate(headers, start=1):
#     ws.cell(row=1, column=col, value=header)

# # Заполняем таблицу данными
# for row, ship in enumerate(ship_data, start=2):
#     for col, header in enumerate(headers, start=1):
#         ws.cell(row=row, column=col, value=ship.get(header, ''))

# # Сохраняем книгу Excel
# wb.save('ships_info.xlsx')